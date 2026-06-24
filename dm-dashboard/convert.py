"""
DM施策別進捗一覧 → JSON 変換スクリプト
使い方: python convert.py <xlsmファイルパス>
例:    python convert.py "DM施策別進捗一覧6.xlsm"

出力: data/dm_data.json （上書き更新）
"""

import sys
import json
import re
from pathlib import Path
from openpyxl import load_workbook


# ── カテゴリ正規化ルール ──────────────────────────────
# 新しい施策が増えた場合はここに追加してください
def normalize_category(promo_name: str) -> str:
    n = promo_name
    if re.search(r'お誕生日|誕生日', n):         return 'お誕生日DM'
    if re.search(r'TRS下取.*WOW',   n):           return 'TRS下取WOW'
    if re.search(r'TRS下取新規',    n):           return 'TRS下取新規'
    if re.search(r'TRS下取',        n):           return 'TRS下取定期'
    if re.search(r'TRSキャンペーン|TRSキャン', n): return 'TRSキャンペーン'
    if re.search(r'RAH買い替え',    n):           return 'RAH買い替え'
    if re.search(r'RAH',            n):           return 'RAHシリーズ'
    if re.search(r'CCHシリーズ',    n):           return 'CCHシリーズ'
    if re.search(r'CCH買い替え|CCH買替', n):      return 'CCH買い替え'
    if re.search(r'CCH',            n):           return 'CCH'
    if re.search(r'TR7買い替え|TR7買替', n):      return 'TR7買い替え'
    if re.search(r'TR7',            n):           return 'TR7'
    if re.search(r'シニア',         n):           return 'シニア訴求'
    return 'その他'


def is_test(promo_name: str) -> bool:
    return 'テスト' in promo_name or 'test' in promo_name.lower()


# ── シート列レイアウト定義 ────────────────────────────
# (promo_col, send_col, resp_col, rev_col, ord_col, p2_col, p3_col, pr_col, media_col)
# ヘッダー行を見て列がずれていたら修正してください
SHEET_LAYOUTS = {
    # 月キー: (シート名, promo列, 発送数列, RR列, 受注金額列, 受注数列, P2列, P3列, PR列, 媒体費列)
    '6':  ('6月詳細',   5, 6, 7, 8, 9,10,11,12,13),
    '7':  ('7月詳細',   5, 6, 7, 8, 9,10,11,12,13),
    '8':  ('8月詳細',   7, 8, 9,10,11,12,13,14,15),
    '9':  ('9月詳細',  12,13,14,15,16,17,18,19,20),
    '10': ('10月詳細', 12,13,14,15,16,17,18,19,20),
    '11': ('11月詳細', 12,13,14,15,16,17,18,19,20),
    '12': ('12月詳細', 13,14,15,16,17,18,19,20,21),
    '1':  ('1月詳細',  13,14,15,16,17,18,19,20,21),
    '2':  ('2月詳細',  13,14,15,16,17,18,19,20,21),
    '3':  ('3月詳細',  13,14,15,16,17,18,19,20,21),
    '4':  ('4月詳細',  13,14,15,16,17,18,19,20,21),
    # 新しい月を追加する場合は以下に追記:
    # '5': ('5月詳細', 13,14,15,16,17,18,19,20,21),
}


def extract_sheet(ws, promo_col, send_col, resp_col, rev_col, ord_col,
                  p2_col, p3_col, pr_col, media_col):
    rows = list(ws.iter_rows(max_row=300, values_only=True))
    promos = []
    for row in rows[2:]:
        if not row or row[1] != 1:
            continue
        try:
            pname = str(row[promo_col]).strip() if row[promo_col] else ''
            if not pname or pname == 'None':
                continue
            send = float(row[send_col]) if row[send_col] else 0
            if send <= 0:
                continue
            promos.append({
                'promo':         pname,
                'category':      normalize_category(pname),
                'is_test':       is_test(pname),
                'send_count':    int(send),
                'response_rate': round(float(row[resp_col]  or 0) * 100, 4),
                'revenue':       round(float(row[rev_col]   or 0)),
                'orders':        int(float(row[ord_col]     or 0)),
                'p2':            round(float(row[p2_col]    or 0)),
                'p3':            round(float(row[p3_col]    or 0)),
                'pr':            round(float(row[pr_col]    or 0), 4),
                'media_cost':    round(float(row[media_col] or 0)),
            })
        except Exception:
            pass
    return promos



# ── 商品別データ抽出 ────────────────────────────────
PROD_LAYOUTS = {
    '6':  ('6月詳細',   5, 6, 8, 9,11),
    '7':  ('7月詳細',   5, 6, 8, 9,11),
    '8':  ('8月詳細',   7, 8,10,11,13),
    '9':  ('9月詳細',  12,13,15,16,18),
    '10': ('10月詳細', 12,13,15,16,18),
    '11': ('11月詳細', 12,13,15,16,18),
    '12': ('12月詳細', 13,14,16,17,19),
    '1':  ('1月詳細',  13,14,16,17,19),
    '2':  ('2月詳細',  13,14,16,17,19),
    '3':  ('3月詳細',  13,14,16,17,19),
    '4':  ('4月詳細',  13,14,16,17,19),
}

PROD_SKIP = {'TOTAL','None','','商品ｸﾞﾙｰﾌﾟ','発送数','ﾌﾟﾛﾓｰｼｮﾝ名','各施策商品ごと成績','集計'}

def extract_products(ws, promo_col, prod_col, rev_col, ord_col, p3_col):
    rows = list(ws.iter_rows(max_row=700, values_only=True))
    results = []
    cur = None
    for row in rows:
        if not row: continue
        if row[1] == 1 and row[promo_col]:
            pname = str(row[promo_col]).strip()
            if not pname or pname == 'None': continue
            cur = pname
            prod = str(row[prod_col]).strip() if row[prod_col] else ''
            if prod and prod not in PROD_SKIP and not prod.replace('.','').replace('-','').lstrip('-').isdigit():
                try:
                    results.append({'promo': cur, 'product': prod, 'category': normalize_category(cur),
                        'revenue': round(float(row[rev_col] or 0)),
                        'orders': int(float(row[ord_col] or 0)),
                        'p3': round(float(row[p3_col] or 0))})
                except: pass
        elif cur and row[prod_col]:
            prod = str(row[prod_col]).strip()
            if prod in PROD_SKIP or not prod: continue
            if prod[0].isdigit(): continue
            try:
                results.append({'promo': cur, 'product': prod, 'category': normalize_category(cur),
                    'revenue': round(float(row[rev_col] or 0)),
                    'orders': int(float(row[ord_col] or 0)),
                    'p3': round(float(row[p3_col] or 0))})
            except: pass
    return results

def main():
    if len(sys.argv) < 2:
        print('使い方: python convert.py <xlsmファイルパス>')
        sys.exit(1)

    xlsm_path = sys.argv[1]
    print(f'読み込み中: {xlsm_path}')
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)

    all_data = {}
    for month_key, (sheet_name, *cols) in SHEET_LAYOUTS.items():
        if sheet_name not in wb.sheetnames:
            print(f'  スキップ: {sheet_name} (シートなし)')
            continue
        ws = wb[sheet_name]
        promos = extract_sheet(ws, *cols)
        all_data[month_key] = promos
        total_p3 = sum(p['p3'] for p in promos)
        print(f'  {sheet_name}: {len(promos)}施策  累計P3={total_p3:>12,.0f}円')

    out_path = Path(__file__).parent / 'data' / 'dm_data.json'
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    # 商品別データも抽出
    prod_data = {}
    for month_key, (sheet_name, *cols) in PROD_LAYOUTS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        prods = extract_products(ws, *cols)
        prod_data[month_key] = prods
        print(f'    商品データ {sheet_name}: {len(prods)}件')

    prod_path = Path(__file__).parent / 'data' / 'dm_products.json'
    with open(prod_path, 'w', encoding='utf-8') as f:
        json.dump(prod_data, f, ensure_ascii=False, indent=2)

    print(f'\n✅ 出力完了: {out_path}')
    print(f'✅ 商品データ: {prod_path}')
    print('次のステップ: git add data/ && git commit -m "update data" && git push')


if __name__ == '__main__':
    main()
